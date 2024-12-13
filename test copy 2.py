from openpyxl import load_workbook
import requests
import time
from datetime import datetime


def get_address_info(address, api_key):
    """
    Get address information from Kakao API
    Returns tuple of empty strings if address not found
    """
    url = "https://dapi.kakao.com/v2/local/search/address.json"
    headers = {"Authorization": f"KakaoAK {api_key}"}
    params = {"query": address}

    try:
        response = requests.get(url, headers=headers, params=params)
        response.raise_for_status()
        data = response.json()

        # Check if we have any results
        if data.get("documents") and data["documents"][0].get("road_address"):
            result = data["documents"][0]
            return (
                result["road_address"].get("zone_no", ""),
                result.get("address", {}).get("address_name", ""),
                result["road_address"].get("address_name", ""),
            )
        return ("", "", "")  # Return empty strings if no result found

    except Exception as e:
        print(f"Error processing address '{address}': {e}")
        return ("", "", "")  # Return empty strings on error


def process_excel_file(input_file, api_key):
    """
    Process Excel file and add address information
    """
    try:
        # Load the workbook and get active sheet
        wb = load_workbook(input_file)
        ws = wb.active

        # Find the column index for '사업장 주소'
        address_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "사업장 주소":
                address_col = col
                break

        if not address_col:
            print("Could not find '사업장 주소' column")
            return

        # Add new column headers
        last_col = ws.max_column
        ws.cell(row=1, column=last_col + 1, value="우편번호")
        ws.cell(row=1, column=last_col + 2, value="지번주소")
        ws.cell(row=1, column=last_col + 3, value="도로명주소")

        # Process each row
        total_rows = ws.max_row
        print(f"Total rows to process: {total_rows}")

        for row in range(2, total_rows + 1):
            address = ws.cell(row=row, column=address_col).value

            # Skip empty addresses
            if not address:
                ws.cell(row=row, column=last_col + 1, value="")
                ws.cell(row=row, column=last_col + 2, value="")
                ws.cell(row=row, column=last_col + 3, value="")
                continue

            # Get address information
            zipcode, jibun, road = get_address_info(address, api_key)

            # Update the worksheet (values will be empty strings if no result found)
            ws.cell(row=row, column=last_col + 1, value=zipcode)
            ws.cell(row=row, column=last_col + 2, value=jibun)
            ws.cell(row=row, column=last_col + 3, value=road)

            # Print progress
            if row % 10 == 0:
                print(f"Processed {row}/{total_rows} rows")

        # Generate output filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"processed_addresses_{timestamp}.xlsx"

        # Save the final file
        wb.save(output_file)
        print(f"Processing complete! File saved as: {output_file}")

    except Exception as e:
        print(f"Error: {e}")


# Example usage
if __name__ == "__main__":
    INPUT_FILE = "/Users/jasur/Projects/sabzi/sabzi_server/addresses.xlsx"  # Replace with your Excel file path
    KAKAO_API_KEY = "6b9ae363528a83504b6f09bab81b4b35"  # Replace with your Kakao API key

    process_excel_file(INPUT_FILE, KAKAO_API_KEY)
